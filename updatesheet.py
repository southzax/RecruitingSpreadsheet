from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
import csv
from csv import reader
from openpyxl.utils import FORMULAE
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import FORMULAE
from openpyxl.utils import quote_sheetname

#TODO:  Upload sheet to Teams, make changes, redownlod, rename and check all the things.


# Create a fresh sheet each time -- find a way to save info from previous sheets
workbook = Workbook()
pw = load_workbook(filename="PreviousWB.xlsx", data_only=True)
dateTimeObj = datetime.now()
timestamp = dateTimeObj.strftime("%d-%b-%Y (%H:%M:%S.%f)")
heading_font = Font(bold=True, size=15)
heading_row_font = Font(bold=True)
center_aligned_text = Alignment(horizontal='center')

status_default = 'New'
initials_default = 'JR'
notes_default = ''

# Style Variables
#TODO:  Style cells!


# Create a new sheet in the workbook.
def new_sheet(sheetname, tabname, index):
    sheetname = workbook.create_sheet(tabname, index)

def format_job_sheet(tabname):
        sheet = workbook[tabname]
        #sheet.add_data_validation(sdv)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 5
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 30
        sheet.column_dimensions['G'].width = 50
        for cell in sheet["1:1"]:
            cell.font = heading_row_font
        #for cell in sheet["A"]:
            #sdv.add(cell)


def format_summary_sheet(tabname):
        sheet = workbook[tabname]
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['E'].width = 40
        sheet.row_dimensions[1].height = 23
        sheet.row_dimensions[4].height = 23
        sheet.row_dimensions[19].height = 23
        sheet["A1"].font = heading_font
        sheet["A4"].font = heading_font
        sheet["B4"].font = heading_font
        sheet["A19"].font = heading_font


# Makes headers for each sheet.
def create_headers(tabname):
        sheet = workbook[tabname]
        sheet["A1"] = "Status"
        sheet["B1"] = "Initials"
        sheet["C1"] = "Submission Date"
        sheet["D1"] = "Applicant Name"
        sheet["E1"] = "Phone Number"
        sheet["F1"] = "Email"
        sheet["G1"] = "Notes"
        sheet["H1"] = "Date Appended"
        sheet["J1"] = "Source"


# Append previous values from each job listing sheet.
def rewrite_previous_values(tabname):
    oldsheet = pw[tabname]
    newsheet = workbook[tabname]
    for row in oldsheet.iter_rows(min_row=2, max_row=1000, values_only=True):
        try:
            if row[3] == None:
                continue
            else:
                newsheet.append(row)
        except IndexError:
            break


# Appends new values from a Zip Recruiter CSV to the appropriate sheet.
def update_sheet_zr(csvfile, tabname):
    filename = csvfile
    sheet = workbook[tabname]
    oldsheet = pw[tabname]
    names = []
    emails = []

    try:
        for row in oldsheet.iter_rows(min_row=2, max_row=1000, values_only=True):
            names.append(row[3])
    except IndexError:
        pass

    try:
        for row in oldsheet.iter_rows(min_row=2, max_row=1000, values_only=True):
            emails.append(row[5])
    except IndexError:
        pass

    with open(filename) as f:
        readcsv = reader(f)
        csvinfo = list(readcsv)
        applicants = csvinfo[1:]
        applist = []
        for row in applicants:
            if (row[4] in names and row[5] in emails):
                continue
            else:
                newrow=[]
                newrow.append(status_default)
                newrow.append(initials_default)
                newrow.append(row[3])
                newrow.append(row[4])
                newrow.append(row[6])
                newrow.append(row[5])
                newrow.append(notes_default)
                newrow.append(timestamp)
                newrow.append("ZR")
                sheet.append(newrow)

#TODO:  Find a way to embed resumes!



# Appends new values from an Indeed CSV to the appropriate sheet
def update_sheet_indeed(csvfile, tabname):
    filename = csvfile
    sheet = workbook[tabname]
    oldsheet = pw[tabname]
    names = []
    emails = []

    #TODO:  Find a way to grab the highest Candidate No. on the current sheet.

    try:
        for row in oldsheet.iter_rows(min_row=2, max_row=1000, values_only=True):
            names.append(row[3])
    except IndexError:
        pass

    try:
        for row in oldsheet.iter_rows(min_row=2, max_row=1000, values_only=True):
            emails.append(row[5])
    except IndexError:
        pass

    with open(filename, encoding="utf16") as f:
        readcsv = reader(f, delimiter = '\t')
        csvinfo = list(readcsv)
        applicants = csvinfo[1:]
        applist = []
        for row in applicants:
            if (row[0] in names and row[1] in emails):
                continue
            else:
                newrow=[]
                newrow.append(status_default)
                newrow.append(initials_default)
                newrow.append(row[9])
                newrow.append(row[0])
                newrow.append(row[2])
                newrow.append(row[1])
                newrow.append(notes_default)
                newrow.append(timestamp)
                newrow.append("Indeed")
                #TODO:  Append Candidate No. Here
                sheet.append(newrow)
#TODO:  Find a way to embed resumes!



# Fill in the summary sheet
def create_summary():
    sheet = workbook['Summary']
    sheet["A1"] = "Last Updated:"
    sheet["B1"] = timestamp
    sheet["A4"] = "Applicant Status:"
    sheet["B4"] = "Description:"

    sheet["A5"] = "New"
    sheet["B5"] = "Just imported, not yet reviewed."

    sheet["A6"] = "Reviewed"
    sheet["B6"] = "Reviewed by a HIHR TM, no decision made."

    sheet["A7"] = "WC: Phone Screen"
    sheet["B7"] = "\'Waiting on Client\' for approval for phone screening."

    sheet["A8"] = "Ready to Screen"
    sheet["B8"] = "Client has approved phone screen."

    sheet["A9"] = "WC: Approval to Schedule"
    sheet["B9"] = "\'Waiting on Client\' for approval to schedule an interview."

    sheet["A10"] = "Ready to Schedule"
    sheet["B10"] = "Client approved, ready to call to schedule interview."

    sheet["A11"] = "WC: Interview Scheduling"
    sheet["B11"] = "\'Waiting on Client\' for interview time."

    sheet["A12"] = "Interview Scheduled"
    sheet["B12"] = "This applicant has been scheduled for an interview."

    sheet["A13"] = "WC: Interview Decision"
    sheet["B13"] = "\'Waiting on Client\' for a final decision post-interview."

    sheet["A14"] = "Accepted for Position"
    sheet["B14"] = "Client has accepted applicant, job offer is pending."

    sheet["A15"] = "Turn Down - Personal"
    sheet["B15"] = "Applicant should be turned down and personally notified."

    sheet["A16"] = "Complete (TD Auto)"
    sheet["B16"] = "Applicant can be turned down in Indeed or Zip Recruiter and auto-notified."

    sheet["A19"] = "Open Positions:"
    sheet["A20"] = "Coming Soon!"



#TODO:  Find a way to delete all "Complete" applicants.
#ABOVE:  Maybe not yet.  Until I can automate that, it's nice to be able to
    #Manually delete them so I know I've done it.

#TODO:  Include a summary of status types (definition of each stage)
# !!UPDATE!! this when we change our job listings!!!
    #sheet["C4"] = ((len(ad_sheet["A"])) - 1)
    #sheet["C5"] = ((len(dig_sheet["A"])) - 1)
    #sheet["C6"] = ((len(app_sheet["A"])) - 1)
    #sheet["C7"] = ((len(car_sheet["A"])) - 1)


# def update_summary():



#TODO:  quick summary of when we need their approval -- checkboxes, save changes
#TODO:  (if possible, a summary of number of candidates at each stage)


#TODO:  I should be able to call a series of function on a single thing, right?
# Find a way to minimize the amount of updating when we change listings






# !!UPDATE!! this when we change our job listings!!!
new_sheet('summary_sheet', "Summary", 0)
new_sheet('ad_sheet', "Office Administrator - ESWA", 1)
new_sheet('dig_sheet', "Digital Marketer - IBS", 2)
new_sheet('app_sheet', "Salon Apprentice - BRANCH", 3)
new_sheet('car_sheet', "Lead Carpenter - Balla", 4)

create_headers('Office Administrator - ESWA')
create_headers('Digital Marketer - IBS')
create_headers('Salon Apprentice - BRANCH')
create_headers('Lead Carpenter - Balla')

format_job_sheet('Office Administrator - ESWA')
format_job_sheet('Digital Marketer - IBS')
format_job_sheet('Salon Apprentice - BRANCH')
format_job_sheet('Lead Carpenter - Balla')

format_summary_sheet('Summary')


rewrite_previous_values('Office Administrator - ESWA')
rewrite_previous_values('Digital Marketer - IBS')
rewrite_previous_values('Salon Apprentice - BRANCH')
rewrite_previous_values('Lead Carpenter - Balla')


update_sheet_zr('officeadminzip.csv', 'Office Administrator - ESWA')
update_sheet_zr('digmarzip.csv', 'Digital Marketer - IBS')
update_sheet_zr('leadcarzip.csv', 'Lead Carpenter - Balla')


update_sheet_indeed('officeadminindeed.csv', 'Office Administrator - ESWA')
update_sheet_indeed('leadcarindeed.csv', 'Lead Carpenter - Balla')


create_summary()
#update_summary()




workbook.save(filename=(str("CurrentApplicantList" + timestamp + ".xlsx")))
